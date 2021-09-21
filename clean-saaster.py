# import spreadsheet library 

import openpyxl


def main():
    fp = open('saastr.txt', 'r')  
    fw = open('saastr.txt', 'w')  
    
    line = fp.readline()
    i = 2
    j = 1
    firstname=""
    lastname=""
    role=""
    company=""
    x = 0
    y= 0
    # Create & Initialize sheet 
    book = openpyxl.Workbook()
    book.create_sheet('saastr_attendees')
    sheet = book.get_sheet_by_name('saastr_attendees')
    sheet.cell(row=1, column=1).value = "First Name"
    sheet.cell(row=1, column=2).value = "Last Name"
    sheet.cell(row=1, column=3).value = "Role"
    sheet.cell(row=1, column=4).value = "Company"
    sheet.cell(row=1, column=5).value = "Linkedin"

    counter = 0
   # Loop through text file 
    while line:
        if (len(line)>2):
            if(counter == 0):
                y = line.find(" ")
                x = 0
            else:
                x = line.find(",")
                y = line.find(" ")
                  
            if (x == 0 ):
                sheet.cell(row=i, column=1).value = line[0:y-1]
                sheet.cell(row=i, column=2).value = line[y+1:len(line)-1]
                book.save('saastr_attendees.xlsx')
                i = i+1
            else:
                sheet.cell(row=j, column=3).value = line[0:y-1]
                sheet.cell(row=j, column=4).value = line[y+1:len(line)-1]
                book.save('saastr_attendees.xlsx')
                j = j+1
                    
        else:
            line =""
            fw.writeline(line)
        line = fp.readline()    


# Call main method
if __name__ == '__main__':
    main()