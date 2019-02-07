# import spreadsheet library 

import openpyxl


def main():
    fp = open('attendees.txt', 'r')  
    
    line = fp.readline()
    i = 2
    j = 1

    # Create & Initialize sheet 
    book = openpyxl.Workbook()
    book.create_sheet('Ines')
    sheet = book.get_sheet_by_name('Ines')
    sheet.cell(row=1, column=1).value = "Full Name"
    sheet.cell(row=1, column=2).value = "Role"
    sheet.cell(row=1, column=3).value = "Company"

   # Loop through text file 
    while line:
        if line.find("-"):
            sheet.cell(row=i, column=j).value = line
            book.save('Ines.xlsx')
            j = j+1    
        else:
            i += 1
            j = 1  

        line = fp.readline()    


# Call main method
if __name__ == '__main__':
    main()