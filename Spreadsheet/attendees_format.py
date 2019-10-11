# import spreadsheet library 

import openpyxl


def main():
    fp = open('mentors.txt', 'r')  
    
    line = fp.readline()
    i = 2
    j = 1

    # Create & Initialize sheet 
    book = openpyxl.Workbook()
    book.create_sheet('Plato-Mentors')
    sheet = book.get_sheet_by_name('Plato-Mentors')
    sheet.cell(row=1, column=1).value = "First Name"
    sheet.cell(row=1, column=2).value = "Last Name"
    sheet.cell(row=1, column=3).value = "Role"
    sheet.cell(row=1, column=4).value = "Plato Network"
    sheet.cell(row=1, column=5).value = "Company Size"

    counter = 0
   # Loop through text file 
    while line:
        if counter<=3:
            sheet.cell(row=i, column=j).value = line
            book.save('Plato-Mentors.xlsx')
            j = j+1    
            counter = counter+1
        else:
            sheet.cell(row=i, column=5).value = "100-1000"
            i += 1
            j = 2  
            
            sheet.cell(row=i, column=1).value = line
            book.save('Plato-Mentors.xlsx')
            counter = 1

        line = fp.readline()    


# Call main method
if __name__ == '__main__':
    main()