import openpyxl
import time     
from selenium import webdriver
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.webdriver.common.keys import Keys



PATH = "/home/azzen/Desktop/Marketer/automation/chromedriver"
options = webdriver.ChromeOptions()
options.add_argument("--disable-notifications")
options.add_argument("--headless")  
     
    # to open chrome webbrowser and maximize the window
driver = webdriver.Chrome(executable_path=PATH, chrome_options=options)
driver.maximize_window()
book = openpyxl.load_workbook('/home/azzen/Desktop/Marketer/automation/Marketing-Toolchain/Spreadsheet/Plato-Mentors.xlsx')
sheet = book.get_sheet_by_name('Plato-Mentors')
sheet.cell(row=1, column=6).value = "Linkedin"
for i in range(3,sheet.max_row+1):
    first_name= sheet.cell(row=i, column=1).value.lower()
    last_name= sheet.cell(row=i, column=2).value.lower()
    link = first_name+"-"+last_name
    driver.get("https://community.platohq.com/mentors/"+link)
    driver.implicitly_wait(20)
    test = driver.find_element_by_class_name("mentor-profile-info")
    chaine = test.find_element_by_css_selector('a').get_attribute('href')
    sheet.cell(row=i, column=6).value = chaine 
    book.save('Plato-Mentors.xlsx')

    

