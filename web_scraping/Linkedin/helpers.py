import openpyxl
import time     
from selenium import webdriver
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.webdriver.common.keys import Keys


PATH = "/home/azzen/Desktop/Marketer/automation/chromedriver"
options = webdriver.ChromeOptions()
options.add_argument("--disable-notifications")
options.add_argument("--headless")  
     
    # to open chrome webbrowser and maximize the window
driver = webdriver.Chrome(executable_path=PATH, chrome_options=options)
driver.maximize_window()
     
    #Implicit Wait when element is taking time to load
driver.implicitly_wait(20)    

def auth(email,password):

    account_credentials(email,password)
    # connect to the specific ip address
    driver.get("http://www.linkedin.com")
     
    # Login to Linkedin
    driver.find_element_by_name("session_key").clear()
    driver.find_element_by_name("session_key").send_keys(email)
    driver.find_element_by_name("session_password").clear()
    driver.find_element_by_name("session_password").send_keys(password)
    driver.find_element_by_class_name("sign-in-form__submit-btn").click()
    time.sleep(5)
    



def send_msg_to_one(name,profile_url,msg):
    driver.get(profile_url)
    driver.find_element_by_class_name("pv-s-profile-actions--message").click()  
    driver.find_element_by_xpath("//div[@contenteditable='true']").send_keys("Hey "+name+","+msg)
    driver.find_element_by_class_name("msg-form__send-button").click()

    # Waiting time 
    time.sleep(5) 

def send_one_msg_to_many(generic_msg,spreadsheet):
    path = "/home/azzen/Desktop/Marketer/automation/Marketing-Toolchain/Spreadsheet/"
    book = openpyxl.load_workbook(path+spreadsheet+'.xlsx')
    sheet = book.get_sheet_by_name(spreadsheet)

    for i in range(3,sheet.max_row+1):
        profile_link= sheet.cell(row=i, column=5).value
        msg = sheet.cell(row=i, column=6).value+generic_msg
        driver.get(profile_link)
        driver.find_element_by_class_name("pv-s-profile-actions--message").click()  
        driver.find_element_by_xpath("//div[@contenteditable='true']").send_keys(msg)
        driver.find_element_by_class_name("msg-form__send-button").click()

        # Waiting time 
        time.sleep(5) 
        


def send_many_msgs_to_many(spreadsheet):
    for i in range(2,1000):
        send_one_msg_to_many(msg,spreadsheet+"_"+i)
    

def account_credentials(email,password):
    email_address = email
    pwd = password
    tab = [email,pwd]
    return tab

def save_scr(link,source):
    driver.get(link)
    driver.save_screenshot(source)

def save_scr_list(path,spreadsheet):
    book = openpyxl.load_workbook(path+spreadsheet+".xlsx")  
    sheet = book.get_sheet_by_name(spreadsheet)

    for i in range(3,sheet.max_row+1):
        link= sheet.cell(row=i, column=1).value
        source = sheet.cell(row=i, column=2).value
        save_scr(link,source)

def connect_second_prospect(path,spreadsheet):
    book = openpyxl.load_workbook(path+spreadsheet+".xlsx")  
    sheet = book.get_sheet_by_name(spreadsheet)

    for i in range(3,sheet.max_row+1):
        profile_link= sheet.cell(row=i, column=1).value
        msg = sheet.cell(row=i, column=2).value
        driver.get(profile_link)
        driver.find_element_by_class_name("artdeco-button__text").click()  
        driver.find_element_by_xpath("//*[@class='mr1 artdeco-button artdeco-button--muted artdeco-button--3 artdeco-button--secondary ember-view']").send_keys(msg)
        
            
    